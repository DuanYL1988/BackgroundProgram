import openpyxl
import DBUtil
# 爬虫数据写入的文件路径
FULL_PATH = "D:\\Project\\00_Markdown\\80_VBA\\CrawlerData.xlsx"
DDL_PATH = "D:\Projects\Tools\VBA\DDL.xlsx"

def writeData(sheetName, dataList, codeMap):
    workbook = openpyxl.load_workbook(filename=FULL_PATH, read_only=False)
    dataSheet = workbook[sheetName]
    # 数据登录
    for row_index, data in enumerate(dataList, start=1):
        for col_index, attr in enumerate(data, start=1):
            if attr == "detalUrl" :
                continue
            # 第一行写入标题
            if row_index == 1 :
                dataSheet.cell(row=row_index, column=col_index).value = attr
            # 从第二行写入数据
            strText = data[attr]
            if isinstance(strText, list):
                strText = ",".join(strText)
            dataSheet.cell(row=row_index+1, column=col_index).value = strText
    # Code登录
    codeSheet = workbook["CODE_LIST"]
    rowNo = 2
    for row_index, key in enumerate(codeMap, start=1):
        data = codeMap[key]
        if data is None:
            continue
        for col_index, attr in enumerate(data, start=1):
            codeSheet.cell(row=rowNo, column=2).value = attr
            codeSheet.cell(row=rowNo, column=3).value = ""
            img = data[attr].get("img")
            # print(rowNo, attr, img)
            if img is not None :
                codeSheet.cell(row=rowNo, column=4).value = img
            rowNo = rowNo+1
    workbook.save(FULL_PATH)

def getColumnInfoFromSheet(fullExcelPath, sheetName, range):
    # 取得工作簿
    workbook = openpyxl.load_workbook(fullExcelPath, data_only=True)
    sheet = workbook[sheetName]
    dataList = []
    for row in sheet.iter_rows(min_row=range["startRow"],max_row=sheet.max_row,min_col=range["startCol"],max_col=range["endCol"], values_only=True):
        if row[12] is not None:
            dataList.append(row)
    return dataList

def getDataFromAllSheet(fullExcelPath, cells, startRow, listProp, loopCell):
    '''
    获取excel中所有sheet的数据,并返回成字典格式
    {fullExcelPath} : excel文件路径
    {cells} : sheet中固定位置单元格坐标和属性名 ex: {'camelName':'B2','logicName':'B3'}
    {startRow} : 开始循环的行数
    {listProp} : 循环数据属性
    {loopCell} : 循环的每行属性名集合
    '''
    # 取得工作簿
    workbook = openpyxl.load_workbook(fullExcelPath, data_only=True)
    # 取得sheets,只取名的话可以用workbook.sheetnames
    # 通过sheet名取得sheet: workbook['sheet名']
    # 取得活跃的sheet:workbook.active
    dataDir = {}
    for sheet in workbook:
        table = getSingleSheetCells(sheet,cells)
        table[listProp] = getLoopCells(sheet, startRow, loopCell)
        dataDir[sheet.title] = table
    return dataDir

def getLoopCells(sheet,startRow,loopCell):
    row = startRow
    list = []
    # 从指定开始行到结束行
    while row < len(sheet["A"]):
        # 循环列
        col = 1
        obj = {}
        for item in loopCell:
            # 当前列名设空值时跳过
            if("" != item) :
                # table["logicName"] = sheet.cell(row=3, column=2).value
                value = sheet.cell(row=row, column=col).value
                obj[item] = noneToEmpty(value)
                col += 1
        row += 1
        list.append(obj)
    return list

def getSingleSheetCells(sheet,cells):
    """
    sheet:sheet对象
    cells:数据对象以及坐标ex> {'camelName':'B2','logicName':'B3'}
    """
    obj = {}
    for key in cells:
        # 绝对定位单元格 table["logicName"] = sheet['B3'].value
        value = transColumnNameToCamel(sheet[cells[key]].value,True)
        obj[key] = noneToEmpty(value)
    return obj

def transColumnNameToCamel(columnName, firstCharUpperFlag):
    """
    将数据字段命名转为程序驼峰命名
    """
    result = ""
    for breakWord in columnName.split("_"):
        result += str.upper(breakWord[0:1]) + str.lower(breakWord[1:])
    result = result if firstCharUpperFlag else str.lower(breakWord[0:1]) + str[1:]
    return result

def noneToEmpty(str):
    return "" if None == str else str

def outputFile(outputText, encode, fileName):
  with open(fileName, mode="w" , encoding=encode) as f:
    f.write(str(outputText))
  f.close()

def updateDDLFromDB():
    quary = '''SELECT
                UCASE(TABLE_NAME)
                , COLUMN_NAME
                , COLUMN_COMMENT
                , UCASE(MID(COLUMN_TYPE,1,INSTR(COLUMN_TYPE,'(')-1)) AS TYPE
                , MID(LEFT(COLUMN_TYPE,LENGTH(COLUMN_TYPE)-1),INSTR(COLUMN_TYPE,'(')+1) AS LENGTH
            FROM
                information_schema.COLUMNS COLS 
            WHERE
                COLS.TABLE_NAME IN ( 
                    SELECT
                        TABLE_NAME 
                    FROM
                        information_schema.TABLES 
                    WHERE
                        TABLE_SCHEMA = 'duanyl'
                )
            ORDER BY 1'''
    result = DBUtil.doSearchQuery(quary)
    TABLE_COL_MP = {}
    for rowData in result:
        tblNm = rowData[0]
        if tblNm not in TABLE_COL_MP:
            TABLE_COL_MP[tblNm] = []
        TABLE_COL_MP[tblNm].append(rowData[1:])

    # 更新DDL
    workbook = openpyxl.load_workbook(filename=DDL_PATH, read_only=False)
    for tableName in TABLE_COL_MP:
        targetSheet = workbook[tableName]
        startRow = 4
        for columnInfo in TABLE_COL_MP[tableName]:
            print(f'{tableName} -> {columnInfo}')
            targetSheet.cell(row=startRow, column=1).value = columnInfo[0]
            targetSheet.cell(row=startRow, column=2).value = columnInfo[1]
            targetSheet.cell(row=startRow, column=4).value = columnInfo[2]
            targetSheet.cell(row=startRow, column=5).value = columnInfo[3]
            startRow = startRow + 1
    workbook.save(DDL_PATH)

updateDDLFromDB()