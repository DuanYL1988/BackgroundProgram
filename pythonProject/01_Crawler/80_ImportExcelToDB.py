# 通用工具类
from openpyxl import load_workbook
import DBUtil
import CrawlerUtils
import json
# 爬虫数据写入的文件路径
FULL_PATH = "D:\\Project\\00_Markdown\\80_VBA\\DDL.xlsx"
CONFIG_TBL = "TABLE_INFO"
TARGET_TBL_NAME = "FIREEMBLEM_HERO"

'''
将excel内容转为insert sql
'''
def initConfigInfo():
    # 取得工作簿
    # data_only=True,单元格是公式时,取得表示的值
    workbook = load_workbook(FULL_PATH, data_only=True)
    # 需要导入的表字段信息
    # 取得Excel中定义的TableInfo表信息
    tableInfoSheet = workbook[CONFIG_TBL]
    # 表字段集合
    columns = []
    # values_only=True只会返回单元格值,不会返回对象
    for row in tableInfoSheet.iter_rows(min_row=5, values_only=True):
        columns.append(row[0])

    targetTblSheet = workbook[TARGET_TBL_NAME]
    dataList = []
    for row in targetTblSheet.iter_rows(min_row=4):
        data = {}
        # 遍历每一个单元格
        for cell in row:
            value = ""
            if cell.value is not None:
                value = cell.value
            data[columns[cell.column-1]] = value
        dataList.append(data)
    # 将excel内容转为insert sql写入DB
    createInsertQuery(columns, dataList)

def createInsertQuery(columns, dataList):
    insertPart = ""
    for column in columns:
        insertPart += column + ","
    insertPart = insertPart[0:len(insertPart)-1]
    query = f"INSERT INTO {CONFIG_TBL} (TABLE_NAME, {insertPart}) VALUES "
    for data in dataList:
        valuePart = f"'{TARGET_TBL_NAME}'"
        for colKey in columns:
            valuePart += f",'{str(data[colKey])}'"
        valuePart = f"({valuePart}),"
        query = query + valuePart
    query = str(query)
    query = query[0: len(query)-1]
    print(query)

def createJsonDataFromDb(tableName):
    # 取得表的字段信息
    result = DBUtil.doSearch(table=CONFIG_TBL, columns="COL_NAME, COL_CAMEL", condition={"TABLE_NAME" : tableName})
    selectColumn = ""
    dataList = []
    for row in result:
        selectColumn += row[0] + ","
    selectColumn = selectColumn[0: len(selectColumn)-1]   
    targetTblList = DBUtil.doSearch(table=tableName, columns=selectColumn, condition=None)
    for data in targetTblList:
        newData = {}
        for i in range(0, len(result)):
            newData[str(result[i][1])] = data[i]
        dataList.append(newData)
#        print()
    outputJsonFile("./CrawlerDataFile/", f"{tableName}.json", dataList)

def outputJsonFile(path, fileName, dataList):
  with open(path + fileName, mode="w" , encoding="utf-8") as f:
    f.write("const DATA_LIST = [\n")
    for data in dataList:
        f.write("    ," + json.dumps(data, separators=(',', ':')).encode('utf-8').decode('unicode_escape') + "\n")
    f.write("]")
  f.close()

def getAppCodeListMap(application):
    resultList = DBUtil.doSearch("CODE_MASTER", "CODE,NAME,IMG_URL", {"APPLICATION": application})
    with open("./CrawlerDataFile/CodeMaster.json", mode="w" , encoding="utf-8") as f:
        f.write("const CODE_MASTER = {\n")
        for record in resultList:
            newData = {'name':record[1], 'imgUrl':record[2]}
            f.write(f'    ,"{record[0]}": ' + json.dumps(newData, separators=(',', ':')).encode('utf-8').decode('unicode_escape') + "\n")
        f.write("}")
    f.close()
    # 武器
    query = "select concat(skill_category,'_',SKILL_CODE) as skillKey,SKILL_NAME_CN,SKILL_ICON from fireemblem_hero_skill"
    with open("./CrawlerDataFile/SkillMaster.json", mode="w" , encoding="utf-8") as f:
        f.write("const SKILL_MASTER = {\n")
        for record in DBUtil.doSearchQuery(query):
            newData = {'name':record[1], 'imgUrl':record[2]}
            f.write(f'    ,"{record[0]}": ' + json.dumps(newData, separators=(',', ':')).encode('utf-8').decode('unicode_escape') + "\n")
        f.write("}")
    f.close()

outputType = "1"

# 处理区分:1.输出对象表数据到json文件 
if outputType == "1" :
    createJsonDataFromDb(TARGET_TBL_NAME)
# 处理区分:2.取得Code信息到codemaster.json
else:
    getAppCodeListMap("FEH")