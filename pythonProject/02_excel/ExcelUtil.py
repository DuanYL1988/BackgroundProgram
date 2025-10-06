import openpyxl

def getDataFromAllSheet(fullExcelPath, cells, startRow, listProp, loopCell):
    '''
    获取excel中所有sheet的数据,并返回成字典格式
    {fullExcelPath} : excel文件路径
    {cells} : sheet中固定位置单元格坐标和属性名 ex: {'camelName':'B2','logicName':'B3'}
    {startRow} : 开始循环的行数
    {listProp} : 循环数据属性
    {loopCell} : 循环的每行属性名集合
    '''
    # 取得工作簿
    workbook = openpyxl.load_workbook(fullExcelPath, data_only=True)
    # 取得sheets,只取名的话可以用workbook.sheetnames
    # 通过sheet名取得sheet: workbook['sheet名']
    # 取得活跃的sheet:workbook.active
    dataDir = {}
    for sheet in workbook:
        table = getSingleSheetCells(sheet,cells)
        table[listProp] = getLoopCells(sheet, startRow, loopCell)
        dataDir[sheet.title] = table
    return dataDir

def getLoopCells(sheet,startRow,loopCell):
    row = startRow
    list = []
    # 从指定开始行到结束行
    while row < len(sheet["A"]):
        # 循环列
        col = 1
        obj = {}
        for item in loopCell:
            # 当前列名设空值时跳过
            if("" != item) :
                # table["logicName"] = sheet.cell(row=3, column=2).value
                value = sheet.cell(row=row, column=col).value
                obj[item] = noneToEmpty(value)
                col += 1
        row += 1
        list.append(obj)
    return list

def getSingleSheetCells(sheet,cells):
    """
    sheet:sheet对象
    cells:数据对象以及坐标ex> {'camelName':'B2','logicName':'B3'}
    """
    obj = {}
    for key in cells:
        # 绝对定位单元格 table["logicName"] = sheet['B3'].value
        value = transColumnNameToCamel(sheet[cells[key]].value,True)
        obj[key] = noneToEmpty(value)
    return obj

def transColumnNameToCamel(columnName, firstCharUpperFlag):
    """
    将数据字段命名转为程序驼峰命名
    """
    result = ""
    for breakWord in columnName.split("_"):
        result += str.upper(breakWord[0:1]) + str.lower(breakWord[1:])
    result = result if firstCharUpperFlag else str.lower(breakWord[0:1]) + str[1:]
    return result

def noneToEmpty(str):
    return "" if None == str else str

def outputFile(outputText, encode, fileName):
  with open(fileName, mode="w" , encoding=encode) as f:
    f.write(str(outputText))
  f.close()

excelFullPath = "./DDL.xlsx"
coordinateCell = {'camelName':'B2','logicName':'B3'}
listProp = "columns"
loopCell = ['column','name','propertyName','type','length','pk','notnull','unique']#,'inputType','insertable','updateable','fifterable','disabled','listVisable','code','default'
startRow = 6
MODEL_LIST = getDataFromAllSheet(excelFullPath, coordinateCell, startRow, listProp, loopCell)
outputFile(MODEL_LIST, "utf-8", "TableList.json")

print("End!")